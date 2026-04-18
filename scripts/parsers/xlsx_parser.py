"""
xlsx_parser.py — Extract data from Excel spreadsheets (.xlsx) into markdown.

This parser:
  1. Reads the .xlsx file.
  2. Extracts data sheet-by-sheet as markdown tables.
  3. Outputs one .md file per workbook into parsed/docs/.
"""

from pathlib import Path

from openpyxl import load_workbook


def parse_xlsx(
    xlsx_path: Path,
    output_dir: Path,
) -> Path:
    """Parse an Excel spreadsheet and write a markdown file.

    Args:
        xlsx_path: Path to the .xlsx file.
        output_dir: Directory to write the parsed markdown (e.g., parsed/docs/).

    Returns:
        Path to the generated markdown file.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{xlsx_path.stem}.md"

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

    lines: list[str] = []
    title = xlsx_path.stem.replace("_", " ").replace("-", " ").title()
    lines.append(f"# {title}")
    lines.append("")
    lines.append(f"> Parsed from `{xlsx_path.name}`")
    lines.append(f"> Sheets: {', '.join(wb.sheetnames)}")
    lines.append("")
    lines.append("---")
    lines.append("")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## {sheet_name}")
        lines.append("")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("*Empty sheet.*")
            lines.append("")
            continue

        # Find the actual data bounds (skip fully empty rows)
        data_rows: list[list[str]] = []
        max_cols = 0
        for row in rows:
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                data_rows.append(cells)
                max_cols = max(max_cols, len(cells))

        if not data_rows:
            lines.append("*Empty sheet.*")
            lines.append("")
            continue

        # Normalize column count
        for i, row in enumerate(data_rows):
            if len(row) < max_cols:
                data_rows[i] = row + [""] * (max_cols - len(row))

        # First row as header
        headers = data_rows[0]
        lines.append("| " + " | ".join(h.replace("|", "\\|") for h in headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")

        for row in data_rows[1:]:
            cells = [c.replace("|", "\\|").replace("\n", " ") for c in row]
            lines.append("| " + " | ".join(cells) + " |")

        lines.append("")

    wb.close()
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path
